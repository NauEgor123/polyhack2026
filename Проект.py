import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import random
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Для работы с Excel установите: pip install openpyxl")

class ScheduleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Школьное расписание - Вся неделя + Мероприятия по классам")
        self.root.geometry("1400x900")

        self.classes = list(range(1, 12))

        self.teachers = {
            "Математика": ["Иванова М.И.", "Петров В.С.", "Сидорова А.А."],
            "Русский язык": ["Смирнова Е.В.", "Кузнецова Т.П.", "Васильева О.Н."],
            "Литература": ["Смирнова Е.В.", "Кузнецова Т.П.", "Николаева Л.С."],
            "История": ["Козлов Д.Н.", "Михайлов А.В.", "Андреева Н.К."],
            "Обществознание": ["Козлов Д.Н.", "Михайлов А.В.", "Павлова С.Н."],
            "География": ["Морозова Е.В.", "Степанов А.И.", "Алексеева Н.В."],
            "Биология": ["Волкова Т.П.", "Зайцева О.В.", "Соловьева Е.А."],
            "Физика": ["Федоров И.П.", "Николаев А.В.", "Тимофеев С.С."],
            "Химия": ["Федоров И.П.", "Николаев А.В.", "Тимофеев С.С."],
            "Физкультура": ["Соколов А.В.", "Морозов И.П.", "Егоров В.В."],
            "Английский язык": ["Алексеева Н.В.", "Романова Т.И.", "Давыдова Е.С."],
            "Информатика": ["Соколов А.В.", "Громов П.Р.", "Николаев А.В."],
            "Музыка": ["Григорьева О.В.", "Павлова С.Н.", "Орлова Е.В."],
            "ИЗО": ["Волкова Т.П.", "Белова М.А.", "Новикова Л.С."],
            "Окружающий мир": ["Иванова М.И.", "Петрова Н.В.", "Сидорова Е.П."],
            "Труд": ["Николаев А.В.", "Федоров И.П.", "Макаров С.С."]
        }

        self.all_teachers = sorted(list(set(
            teacher for teachers in self.teachers.values() for teacher in teachers
        )))

        self.subjects = {
            'elementary': ["Математика", "Русский язык", "Литература", "Окружающий мир",
                          "Физкультура", "ИЗО", "Музыка", "Труд", "Английский язык"],
            'middle': ["Математика", "Русский язык", "Литература", "История",
                      "Обществознание", "География", "Биология", "Физика",
                      "Физкультура", "Английский язык", "Информатика", "Музыка", "ИЗО"],
            'high': ["Математика", "Русский язык", "Литература", "История",
                    "Обществознание", "География", "Биология", "Физика",
                    "Химия", "Физкультура", "Английский язык", "Информатика"]
        }

        self.lesson_times = [
            "08:30-09:15", "09:25-10:10", "10:20-11:05",
            "11:25-12:10", "12:20-13:05", "13:15-14:00",
            "14:10-14:55", "15:05-15:50"
        ]

        self.days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]

        self.classrooms = [f"Каб.{i}" for i in range(101, 121)]

        self.subject_colors = {
            "Математика": "#FFB6C1",
            "Русский язык": "#98FB98",
            "Литература": "#98FB98",
            "История": "#FFD700",
            "Обществознание": "#FFD700",
            "География": "#87CEEB",
            "Биология": "#90EE90",
            "Физика": "#DDA0DD",
            "Химия": "#DDA0DD",
            "Физкультура": "#FFA07A",
            "Английский язык": "#FFE4B5",
            "Информатика": "#B0C4DE",
            "Музыка": "#E6E6FA",
            "ИЗО": "#E6E6FA",
            "Окружающий мир": "#98FB98",
            "Труд": "#DEB887"
        }

        self.event_colors = {
            "Собрание": "#FF69B4",
            "Кружок": "#9370DB",
            "Секция": "#20B2AA",
            "Факультатив": "#FF8C00",
            "Консультация": "#6495ED",
            "Экскурсия": "#CD853F",
            "Олимпиада": "#DC143C",
            "Репетиция": "#8FBC8F",
            "Другое": "#A9A9A9"
        }

        self.weekly_schedule = {}

        self.events = {day: {class_num: [] for class_num in self.classes} for day in self.days}
        
        self.substitutions = []
        self.classroom_changes = []
        self.lessons_count = {c: 6 for c in self.classes}
        self.teacher_load = {}

        self.event_class_vars = {}  

        self.color_legend = {
            "Предметы": {
                "Точные науки (розовый)": "#FFB6C1",
                "Языки (зеленый)": "#98FB98", 
                "Гуманитарные (золотой)": "#FFD700",
                "Естественные (голубой)": "#87CEEB",
                "Спорт (лососевый)": "#FFA07A",
                "Творчество (лавандовый)": "#E6E6FA",
                "IT (стальной)": "#B0C4DE"
            },
            "Статусы уроков": {
                "✓ - Обычный урок": "#FFFFFF",
                "✅👤 - Замена учителя": "#FFE4B5",
                "✅🏫 - Смена кабинета": "#E0FFFF",
                "✅👥 - Замена + смена": "#DDA0DD"
            },
            "Мероприятия": {
                "Собрание (розовый)": "#FF69B4",
                "Кружок (фиолетовый)": "#9370DB",
                "Секция (зеленый)": "#20B2AA",
                "Факультатив (оранжевый)": "#FF8C00",
                "Консультация (синий)": "#6495ED",
                "Экскурсия (коричневый)": "#CD853F",
                "Олимпиада (красный)": "#DC143C"
            }
        }
        
        self.setup_ui()
        
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X)
        
        title = ttk.Label(title_frame, text="📚 Школьное расписание - Мероприятия по классам", 
                         font=("Arial", 18, "bold"))
        title.pack(side=tk.LEFT)

        self.time_label = ttk.Label(title_frame, text="", font=("Arial", 10))
        self.time_label.pack(side=tk.RIGHT)
        self.update_time()

        control = ttk.LabelFrame(main_frame, text="Управление", padding="10")
        control.pack(fill=tk.X, pady=5)

        lessons_frame = ttk.Frame(control)
        lessons_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(lessons_frame, text="Количество уроков для классов:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        
        grid = ttk.Frame(lessons_frame)
        grid.pack(pady=5)
        
        self.lesson_vars = {}
        row, col = 0, 0
        for class_num in self.classes:
            f = ttk.Frame(grid, relief=tk.RIDGE, borderwidth=1)
            f.grid(row=row, column=col, padx=2, pady=2, sticky="nsew")
            
            ttk.Label(f, text=f"{class_num} кл:").pack(side=tk.LEFT, padx=2)
            var = tk.IntVar(value=6)
            self.lesson_vars[class_num] = var
            ttk.Spinbox(f, from_=1, to=8, textvariable=var, width=3).pack(side=tk.LEFT, padx=2)
            
            col += 1
            if col > 5:
                col = 0
                row += 1

        btn_frame = ttk.Frame(control)
        btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(btn_frame, text="🔄 Сгенерировать ВСЮ неделю", 
                  command=self.generate_week).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📅 Добавить мероприятие", 
                  command=self.add_event_dialog).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📋 Все мероприятия", 
                  command=self.show_all_events).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📄 Сохранить в TXT", 
                  command=self.save_to_txt).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📊 Сохранить в Excel", 
                  command=self.save_to_excel).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📊 Статистика", 
                  command=self.show_teacher_load).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📝 Журнал замен", 
                  command=self.show_substitutions).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="❌ Очистить всё", 
                  command=self.clear_all).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🎨 Легенда цветов", 
                  command=self.show_color_legend).pack(side=tk.LEFT, padx=2)

        ttk.Button(btn_frame, text="👨‍🏫 Расписание учителя", 
                  command=self.show_teacher_schedule).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🔍 Поиск", 
                  command=self.search_dialog).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📊 Статистика классов", 
                  command=self.show_class_stats).pack(side=tk.LEFT, padx=2)

        quick_frame = ttk.Frame(control)
        quick_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(quick_frame, text="Всем 6", 
                  command=lambda: self.set_all_lessons(6)).pack(side=tk.LEFT, padx=2)
        ttk.Button(quick_frame, text="Начальной 5", 
                  command=lambda: self.set_level_lessons(1, 4, 5)).pack(side=tk.LEFT, padx=2)
        ttk.Button(quick_frame, text="Средней 6", 
                  command=lambda: self.set_level_lessons(5, 9, 6)).pack(side=tk.LEFT, padx=2)
        ttk.Button(quick_frame, text="Старшей 7", 
                  command=lambda: self.set_level_lessons(10, 11, 7)).pack(side=tk.LEFT, padx=2)

        sub_panel = ttk.Frame(control)
        sub_panel.pack(fill=tk.X, pady=5)

        teacher_sub_frame = ttk.LabelFrame(sub_panel, text="🔁 Замена учителя", padding="5")
        teacher_sub_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        
        ttk.Label(teacher_sub_frame, text="День:").pack(side=tk.LEFT, padx=2)
        self.sub_day = ttk.Combobox(teacher_sub_frame, values=self.days, width=10, state="readonly")
        self.sub_day.pack(side=tk.LEFT, padx=2)
        self.sub_day.bind('<<ComboboxSelected>>', self.update_teachers_for_sub)
        
        ttk.Label(teacher_sub_frame, text="Класс:").pack(side=tk.LEFT, padx=2)
        self.sub_class = ttk.Combobox(teacher_sub_frame, values=self.classes, width=5, state="readonly")
        self.sub_class.pack(side=tk.LEFT, padx=2)
        self.sub_class.bind('<<ComboboxSelected>>', self.update_teachers_for_sub)
        
        ttk.Label(teacher_sub_frame, text="Урок:").pack(side=tk.LEFT, padx=2)
        self.sub_lesson = ttk.Spinbox(teacher_sub_frame, from_=1, to=8, width=5)
        self.sub_lesson.pack(side=tk.LEFT, padx=2)
        self.sub_lesson.bind('<KeyRelease>', self.update_teachers_for_sub)
        
        ttk.Label(teacher_sub_frame, text="Новый учитель:").pack(side=tk.LEFT, padx=2)
        self.sub_teacher = ttk.Combobox(teacher_sub_frame, width=20)
        self.sub_teacher.pack(side=tk.LEFT, padx=2)
        
        ttk.Button(teacher_sub_frame, text="✏️ Заменить", 
                  command=self.manual_substitute).pack(side=tk.LEFT, padx=2)

        room_sub_frame = ttk.LabelFrame(sub_panel, text="🏫 Замена кабинета", padding="5")
        room_sub_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        
        ttk.Label(room_sub_frame, text="День:").pack(side=tk.LEFT, padx=2)
        self.room_day = ttk.Combobox(room_sub_frame, values=self.days, width=10, state="readonly")
        self.room_day.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(room_sub_frame, text="Класс:").pack(side=tk.LEFT, padx=2)
        self.room_class = ttk.Combobox(room_sub_frame, values=self.classes, width=5, state="readonly")
        self.room_class.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(room_sub_frame, text="Урок:").pack(side=tk.LEFT, padx=2)
        self.room_lesson = ttk.Spinbox(room_sub_frame, from_=1, to=8, width=5)
        self.room_lesson.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(room_sub_frame, text="Новый кабинет:").pack(side=tk.LEFT, padx=2)
        self.room_entry = ttk.Entry(room_sub_frame, width=15)
        self.room_entry.pack(side=tk.LEFT, padx=2)
        
        ttk.Button(room_sub_frame, text="✏️ Сменить", 
                  command=self.manual_room_change).pack(side=tk.LEFT, padx=2)

        warning_label = ttk.Label(control, text="⚠️ Данные НЕ сохраняются автоматически! Используйте кнопки сохранения", 
                                 foreground="red", font=("Arial", 9, "bold"))
        warning_label.pack(pady=5)

        self.day_notebook = ttk.Notebook(main_frame)
        self.day_notebook.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.day_trees = {}  
        self.event_frames = {} 
        
        for day in self.days:
            self.create_day_tab(day)

        self.status = ttk.Label(main_frame, text="Готов к работе. Данные НЕ сохраняются автоматически!", relief=tk.SUNKEN)
        self.status.pack(fill=tk.X, pady=5)

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def on_closing(self):
        """При закрытии окна спрашиваем, сохранять ли"""
        if messagebox.askyesno("Выход", "Сохранить данные перед выходом?"):
            self.save_to_txt()
        self.root.destroy()
    
    def create_day_tab(self, day):
        """Создает вкладку для дня недели с разделом мероприятий"""
        day_frame = ttk.Frame(self.day_notebook)
        self.day_notebook.add(day_frame, text=day)

        paned = ttk.PanedWindow(day_frame, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True)

        top_frame = ttk.LabelFrame(paned, text=f"📚 Расписание уроков - {day}")
        paned.add(top_frame, weight=3)

        legend_frame = ttk.Frame(top_frame)
        legend_frame.pack(fill=tk.X, padx=5, pady=2)
        
        ttk.Label(legend_frame, text="Статусы уроков:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        ttk.Label(legend_frame, text="✓ - Обычный", foreground="green").pack(side=tk.LEFT, padx=5)
        ttk.Label(legend_frame, text="✅👤 - Замена учителя", foreground="orange").pack(side=tk.LEFT, padx=5)
        ttk.Label(legend_frame, text="✅🏫 - Смена кабинета", foreground="blue").pack(side=tk.LEFT, padx=5)
        ttk.Label(legend_frame, text="✅👥 - Замена + смена", foreground="purple").pack(side=tk.LEFT, padx=5)

        class_notebook = ttk.Notebook(top_frame)
        class_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.day_trees[day] = {}

        for class_num in self.classes:
            self.create_class_tab_in_day(day, class_num, class_notebook)

        bottom_frame = ttk.LabelFrame(paned, text=f"🎉 Мероприятия класса", padding="5")
        paned.add(bottom_frame, weight=1)

        select_frame = ttk.Frame(bottom_frame)
        select_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(select_frame, text="Показать мероприятия для класса:").pack(side=tk.LEFT, padx=2)

        self.event_class_vars[day] = tk.StringVar(value="1")
        event_class_combo = ttk.Combobox(select_frame, textvariable=self.event_class_vars[day], 
                                        values=self.classes, width=5, state="readonly")
        event_class_combo.pack(side=tk.LEFT, padx=2)
        event_class_combo.bind('<<ComboboxSelected>>', lambda e, d=day: self.update_events_display(d))
        
        ttk.Button(select_frame, text=f"➕ Добавить мероприятие", 
                  command=lambda d=day: self.add_event_dialog(d, self.event_class_vars[day].get())).pack(side=tk.LEFT, padx=20)
        ttk.Button(select_frame, text="🔄 Обновить", 
                  command=lambda d=day: self.update_events_display(d)).pack(side=tk.LEFT, padx=2)

        event_text = tk.Text(bottom_frame, wrap=tk.WORD, height=6, font=("Arial", 10))
        scroll = ttk.Scrollbar(bottom_frame, orient=tk.VERTICAL, command=event_text.yview)
        event_text.configure(yscrollcommand=scroll.set)
        
        event_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=2)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.event_frames[day] = event_text
    
    def create_class_tab_in_day(self, day, class_num, notebook):
        """Создает вкладку класса внутри дня"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text=f"{class_num} класс")

        if class_num <= 4:
            level = "🏫 Начальная школа"
        elif class_num <= 9:
            level = "📚 Средняя школа"
        else:
            level = "🎓 Старшая школа"
        
        ttk.Label(frame, text=level, font=("Arial", 9, "italic")).pack(pady=2)

        columns = ("№", "Время", "Предмет", "Учитель", "Кабинет", "Статус")
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=8)

        widths = [40, 100, 200, 180, 80, 80]
        for col, width in zip(columns, widths):
            tree.heading(col, text=col)
            tree.column(col, width=width, anchor="center")

        scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.day_trees[day][class_num] = tree
    
    def update_time(self):
        """Обновляет время"""
        now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        self.time_label.config(text=now)
        self.root.after(1000, self.update_time)
    
    def set_all_lessons(self, count):
        """Устанавливает всем классам одинаковое количество уроков"""
        for var in self.lesson_vars.values():
            var.set(count)
        self.status.config(text=f"✅ Всем классам {count} уроков")
    
    def set_level_lessons(self, start, end, count):
        """Устанавливает количество уроков для уровня классов"""
        for class_num in range(start, end + 1):
            if class_num in self.lesson_vars:
                self.lesson_vars[class_num].set(count)
        self.status.config(text=f"✅ {start}-{end} классам {count} уроков")
    
    def clear_all(self):
        """Очищает все данные"""
        if messagebox.askyesno("Очистка", "Вы уверены, что хотите очистить все данные?"):
            self.weekly_schedule = {}
            for day in self.days:
                for class_num in self.classes:
                    tree = self.day_trees[day][class_num]
                    tree.delete(*tree.get_children())
            
            self.events = {day: {class_num: [] for class_num in self.classes} for day in self.days}
            for day in self.days:
                self.update_events_display(day)
            
            self.substitutions = []
            self.classroom_changes = []
            self.teacher_load = {}
            
            self.status.config(text="✅ Все данные очищены")
    
    def get_level(self, class_num):
        """Определяет уровень класса"""
        if class_num <= 4:
            return 'elementary'
        elif class_num <= 9:
            return 'middle'
        else:
            return 'high'
    
    def generate_week(self):
        """Генерирует расписание на ВСЮ неделю (6 разных дней)"""
        try:
            for class_num, var in self.lesson_vars.items():
                self.lessons_count[class_num] = var.get()

            for day in self.days:
                for class_num in self.classes:
                    tree = self.day_trees[day][class_num]
                    tree.delete(*tree.get_children())

            self.weekly_schedule = {}
            self.teacher_load = {}

            random.seed()
            
            for day_index, day in enumerate(self.days):
                day_schedule = {}

                teacher_busy = {}
                
                for class_num in self.classes:
                    level = self.get_level(class_num)
                    subjects = self.subjects[level].copy()

                    for _ in range(day_index * 3 + 5):
                        random.shuffle(subjects)

                    shift = random.randint(1, len(subjects) - 1)
                    subjects = subjects[shift:] + subjects[:shift]
                    
                    schedule = []
                    lessons_count = self.lessons_count[class_num]
                    
                    for i in range(lessons_count):
                        subject = subjects[i % len(subjects)]

                        teachers_list = self.teachers.get(subject, ["Нет учителя"]).copy()

                        chosen_teacher = None

                        random.shuffle(teachers_list)
                        
                        for teacher in teachers_list:
                            busy = False
                            if teacher in teacher_busy:
                                for (c, lesson_i) in teacher_busy[teacher]:
                                    if lesson_i == i:
                                        busy = True
                                        break
                            
                            if not busy:
                                chosen_teacher = teacher
                                break

                        if not chosen_teacher and teachers_list:
                            chosen_teacher = teachers_list[0]
                        
                        if not chosen_teacher:
                            chosen_teacher = "Нет учителя"

                        if chosen_teacher not in teacher_busy:
                            teacher_busy[chosen_teacher] = []
                        teacher_busy[chosen_teacher].append((class_num, i))

                        classroom = random.choice(self.classrooms)
                        
                        lesson = {
                            'number': i + 1,
                            'time': self.lesson_times[i],
                            'subject': subject,
                            'teacher': chosen_teacher,
                            'classroom': classroom,
                            'original_teacher': chosen_teacher,
                            'original_classroom': classroom,
                            'substituted': False,
                            'room_changed': False
                        }
                        
                        schedule.append(lesson)

                        tree = self.day_trees[day][class_num]
                        status = "✓"
                        item = tree.insert("", tk.END, values=(
                            i + 1,
                            self.lesson_times[i],
                            subject,
                            chosen_teacher,
                            classroom,
                            status
                        ))

                        color = self.subject_colors.get(subject, "white")
                        tree.tag_configure(subject, background=color)
                        tree.item(item, tags=(subject,))
                    
                    day_schedule[class_num] = schedule
                
                self.weekly_schedule[day] = day_schedule

            self.teacher_load = {}
            for day in self.days:
                for class_num in self.classes:
                    for lesson in self.weekly_schedule[day][class_num]:
                        teacher = lesson['teacher']
                        if teacher not in self.teacher_load:
                            self.teacher_load[teacher] = 0
                        self.teacher_load[teacher] += 1

            for day in self.days:
                self.update_events_display(day)
            
            self.status.config(text="✅ Расписание на ВСЮ неделю создано! Все 6 дней разные.")
            
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
    
    def add_event_dialog(self, default_day=None, default_class=None):
        """Диалог добавления мероприятия для конкретного класса"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Добавить мероприятие для класса")
        dialog.geometry("500x520")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Добавление мероприятия для класса", font=("Arial", 14, "bold")).pack(pady=10)
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="День:").grid(row=0, column=0, sticky=tk.W, pady=5)
        day_var = tk.StringVar(value=default_day if default_day else "Понедельник")
        day_combo = ttk.Combobox(frame, textvariable=day_var, values=self.days, state="readonly", width=15)
        day_combo.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Класс:").grid(row=1, column=0, sticky=tk.W, pady=5)
        class_var = tk.StringVar(value=default_class if default_class else "1")
        class_combo = ttk.Combobox(frame, textvariable=class_var, values=self.classes, state="readonly", width=5)
        class_combo.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Время начала (часы:минуты):").grid(row=2, column=0, sticky=tk.W, pady=5)
        time_frame = ttk.Frame(frame)
        time_frame.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        
        time_hour = ttk.Entry(time_frame, width=5)
        time_hour.pack(side=tk.LEFT)
        ttk.Label(time_frame, text=":").pack(side=tk.LEFT)
        time_min = ttk.Entry(time_frame, width=5)
        time_min.pack(side=tk.LEFT)
        
        time_hour.insert(0, "15")
        time_min.insert(0, "30")

        ttk.Label(frame, text="(часы 0-23, минуты 0-59)", font=("Arial", 8), foreground="gray").grid(row=2, column=2, sticky=tk.W, padx=5)

        ttk.Label(frame, text="Продолжительность (минут):").grid(row=3, column=0, sticky=tk.W, pady=5)
        duration_var = tk.StringVar(value="45")
        duration_entry = ttk.Entry(frame, textvariable=duration_var, width=10)
        duration_entry.grid(row=3, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Тип мероприятия:").grid(row=4, column=0, sticky=tk.W, pady=5)
        event_types = list(self.event_colors.keys())
        type_var = tk.StringVar(value=event_types[0])
        type_combo = ttk.Combobox(frame, textvariable=type_var, values=event_types, state="readonly", width=20)
        type_combo.grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Название:").grid(row=5, column=0, sticky=tk.W, pady=5)
        name_entry = ttk.Entry(frame, width=40)
        name_entry.grid(row=5, column=1, columnspan=2, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Ответственный:").grid(row=6, column=0, sticky=tk.W, pady=5)
        teacher_var = tk.StringVar()
        teacher_combo = ttk.Combobox(frame, textvariable=teacher_var, values=self.all_teachers, width=30)
        teacher_combo.grid(row=6, column=1, columnspan=2, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Место проведения:").grid(row=7, column=0, sticky=tk.W, pady=5)
        place_entry = ttk.Entry(frame, width=40)
        place_entry.insert(0, "Каб. 101")
        place_entry.grid(row=7, column=1, columnspan=2, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Участники:").grid(row=8, column=0, sticky=tk.W, pady=5)
        participants_entry = ttk.Entry(frame, width=40)
        participants_entry.insert(0, "Учащиеся класса")
        participants_entry.grid(row=8, column=1, columnspan=2, sticky=tk.W, padx=5, pady=5)

        ttk.Label(frame, text="Описание:").grid(row=9, column=0, sticky=tk.W, pady=5)
        desc_text = tk.Text(frame, width=40, height=3)
        desc_text.grid(row=9, column=1, columnspan=2, sticky=tk.W, padx=5, pady=5)
        
        def save_event():
            """Сохраняет мероприятие"""
            day = day_var.get()
            class_num = int(class_var.get())

            hour_str = time_hour.get().strip()
            minute_str = time_min.get().strip()
            
            if not hour_str or not minute_str:
                messagebox.showwarning("Ошибка", "Введите время начала")
                return
            
            try:
                hour = int(hour_str)
                minute = int(minute_str)
                
                if hour < 0 or hour > 23:
                    messagebox.showwarning("Ошибка", "Часы должны быть от 0 до 23")
                    return
                if minute < 0 or minute > 59:
                    messagebox.showwarning("Ошибка", "Минуты должны быть от 0 до 59")
                    return
            except ValueError:
                messagebox.showwarning("Ошибка", "Введите корректное время")
                return
            
            try:
                duration = int(duration_var.get())
                if duration < 1 or duration > 300:
                    messagebox.showwarning("Ошибка", "Продолжительность от 1 до 300 минут")
                    return
            except ValueError:
                messagebox.showwarning("Ошибка", "Введите корректную продолжительность")
                return
            
            total_minutes = hour * 60 + minute + duration
            end_hour = (total_minutes // 60) % 24
            end_minute = total_minutes % 60
            
            time_str = f"{hour:02d}:{minute:02d}-{end_hour:02d}:{end_minute:02d}"
            
            event_type = type_var.get()
            name = name_entry.get().strip()
            teacher = teacher_var.get()
            place = place_entry.get().strip()
            participants = participants_entry.get().strip()
            description = desc_text.get("1.0", tk.END).strip()
            
            if not name:
                messagebox.showwarning("Ошибка", "Введите название мероприятия")
                return
            
            if not teacher:
                messagebox.showwarning("Ошибка", "Выберите ответственного учителя")
                return
            
            if not place:
                messagebox.showwarning("Ошибка", "Введите место проведения")
                return

            event = {
                'time': time_str,
                'type': event_type,
                'name': name,
                'teacher': teacher,
                'place': place,
                'participants': participants if participants else "Учащиеся класса",
                'description': description,
                'class': class_num,
                'datetime': datetime.now().strftime("%Y-%m-%d %H:%M")
            }

            if day not in self.events:
                self.events[day] = {c: [] for c in self.classes}
            if class_num not in self.events[day]:
                self.events[day][class_num] = []
            
            self.events[day][class_num].append(event)

            def time_to_minutes(t):
                try:
                    start = t.split('-')[0]
                    h, m = map(int, start.split(':'))
                    return h * 60 + m
                except:
                    return 0
            
            self.events[day][class_num].sort(key=lambda x: time_to_minutes(x['time']))

            if day in self.event_class_vars:
                self.event_class_vars[day].set(str(class_num))

            self.update_events_display(day)
            
            self.status.config(text=f"✅ Мероприятие '{name}' добавлено для {class_num} класса на {day}")
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="✅ Добавить", command=save_event).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Отмена", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def update_events_display(self, day):
        """Обновляет отображение мероприятий для конкретного дня и выбранного класса"""
        if day not in self.event_frames:
            return

        if day not in self.event_class_vars:
            self.event_class_vars[day] = tk.StringVar(value="1")
        
        try:
            class_num = int(self.event_class_vars[day].get())
        except:
            class_num = 1
            self.event_class_vars[day].set("1")
        
        text_widget = self.event_frames[day]
        text_widget.delete(1.0, tk.END)
 
        if day in self.events and class_num in self.events[day] and self.events[day][class_num]:
            text_widget.insert(tk.END, f"📌 МЕРОПРИЯТИЯ ДЛЯ {class_num} КЛАССА:\n\n", "header")
            for event in self.events[day][class_num]:
                color = self.event_colors.get(event['type'], "#A9A9A9")

                text_widget.insert(tk.END, f"🕒 {event['time']} ", "time")
                text_widget.insert(tk.END, f"[{event['type']}] ", "type")
                text_widget.insert(tk.END, f"{event['name']}\n", "name")
                text_widget.insert(tk.END, f"   👤 {event['teacher']}  ", "teacher")
                text_widget.insert(tk.END, f"📍 {event['place']}\n", "place")
                text_widget.insert(tk.END, f"   👥 {event['participants']}\n", "participants")
                if event['description']:
                    text_widget.insert(tk.END, f"   📝 {event['description']}\n", "desc")
                text_widget.insert(tk.END, "-"*50 + "\n", "separator")

            text_widget.tag_configure("header", font=("Arial", 10, "bold"), foreground="darkblue")
            text_widget.tag_configure("time", font=("Arial", 9, "bold"), foreground="blue")
            text_widget.tag_configure("type", font=("Arial", 9, "bold"), foreground="purple")
            text_widget.tag_configure("name", font=("Arial", 10, "bold"), foreground="darkblue")
            text_widget.tag_configure("teacher", font=("Arial", 9), foreground="green")
            text_widget.tag_configure("place", font=("Arial", 9), foreground="brown")
            text_widget.tag_configure("participants", font=("Arial", 9), foreground="gray")
            text_widget.tag_configure("desc", font=("Arial", 8, "italic"), foreground="gray")
            text_widget.tag_configure("separator", font=("Arial", 8))
        else:
            text_widget.insert(tk.END, f"🎉 Нет мероприятий для {class_num} класса\n")
            text_widget.insert(tk.END, "Нажмите 'Добавить мероприятие' чтобы создать")
    
    def show_color_legend(self):
        """Показывает легенду цветов"""
        win = tk.Toplevel(self.root)
        win.title("Легенда цветов")
        win.geometry("500x600")
        
        main_frame = ttk.Frame(win, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="🎨 Легенда цветовых обозначений", 
                 font=("Arial", 16, "bold")).pack(pady=10)

        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        ttk.Label(scrollable_frame, text="📚 Предметы по категориям", 
                 font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(10,5))
        
        for desc, color in self.color_legend["Предметы"].items():
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, pady=2)
            
            color_label = tk.Label(frame, width=2, height=1, bg=color, relief=tk.RIDGE)
            color_label.pack(side=tk.LEFT, padx=5)
            
            ttk.Label(frame, text=desc).pack(side=tk.LEFT, padx=5)

        ttk.Label(scrollable_frame, text="\n✅ Статусы уроков", 
                 font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(10,5))
        
        for desc, color in self.color_legend["Статусы уроков"].items():
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, pady=2)
            
            if "✓" in desc:
                ttk.Label(frame, text=desc).pack(side=tk.LEFT, padx=5)
            else:
                preview = tk.Label(frame, text="  Пример  ", bg="#FFE4B5", relief=tk.RIDGE)
                preview.pack(side=tk.LEFT, padx=5)
                ttk.Label(frame, text=desc).pack(side=tk.LEFT, padx=5)

        ttk.Label(scrollable_frame, text="\n🎉 Мероприятия", 
                 font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(10,5))
        
        for desc, color in self.color_legend["Мероприятия"].items():
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, pady=2)
            
            color_label = tk.Label(frame, width=2, height=1, bg=color, relief=tk.RIDGE)
            color_label.pack(side=tk.LEFT, padx=5)
            
            ttk.Label(frame, text=desc).pack(side=tk.LEFT, padx=5)

        ttk.Label(scrollable_frame, text="\n📌 Пояснения:", 
                 font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(10,5))
        
        explanations = [
            "• 🔁 - Учитель заменен",
            "• 🏫 - Кабинет изменен",
            "• ✅👤 - Замена учителя",
            "• ✅🏫 - Смена кабинета", 
            "• ✅👥 - Замена учителя и кабинета",
            "• Каждый предмет имеет свой цвет для быстрого поиска"
        ]
        
        for exp in explanations:
            ttk.Label(scrollable_frame, text=exp).pack(anchor=tk.W, pady=1)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def show_teacher_schedule(self):
        """Показывает расписание для выбранного учителя"""
        win = tk.Toplevel(self.root)
        win.title("Расписание учителя")
        win.geometry("1000x700")

        top_frame = ttk.Frame(win)
        top_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(top_frame, text="Выберите учителя:", font=("Arial", 11, "bold")).pack(side=tk.LEFT, padx=5)
        teacher_var = tk.StringVar()
        teacher_combo = ttk.Combobox(top_frame, textvariable=teacher_var, 
                                     values=self.all_teachers, width=30, state="readonly")
        teacher_combo.pack(side=tk.LEFT, padx=5)
        teacher_combo.bind('<<ComboboxSelected>>', lambda e: update_schedule())

        stats_frame = ttk.Frame(top_frame)
        stats_frame.pack(side=tk.RIGHT, padx=10)
        
        total_label = ttk.Label(stats_frame, text="Всего уроков: 0", font=("Arial", 10, "bold"))
        total_label.pack(side=tk.LEFT, padx=10)
        
        sub_label = ttk.Label(stats_frame, text="Замен: 0", font=("Arial", 10))
        sub_label.pack(side=tk.LEFT, padx=10)

        columns = ("День", "Класс", "Урок", "Время", "Предмет", "Кабинет", "Статус")
        tree = ttk.Treeview(win, columns=columns, show="headings", height=25)

        col_widths = [100, 60, 50, 100, 200, 80, 80]
        for col, width in zip(columns, col_widths):
            tree.heading(col, text=col)
            tree.column(col, width=width, anchor="center")

        scroll_y = ttk.Scrollbar(win, orient=tk.VERTICAL, command=tree.yview)
        scroll_x = ttk.Scrollbar(win, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=5)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        def update_schedule():
            teacher = teacher_var.get()
            if not teacher:
                messagebox.showwarning("Ошибка", "Выберите учителя")
                return
            
            tree.delete(*tree.get_children())
            total_lessons = 0
            total_subs = 0

            for day in self.days:
                day_lessons = []
                for class_num in self.classes:
                    if day in self.weekly_schedule and class_num in self.weekly_schedule[day]:
                        for lesson in self.weekly_schedule[day][class_num]:
                            if lesson['teacher'] == teacher:
                                total_lessons += 1
                                if lesson['substituted']:
                                    total_subs += 1
                                
                                status = ""
                                if lesson['substituted'] and lesson['room_changed']:
                                    status = "✅👥"
                                elif lesson['substituted']:
                                    status = "✅👤"
                                elif lesson['room_changed']:
                                    status = "✅🏫"
                                
                                day_lessons.append((
                                    day, f"{class_num} класс", lesson['number'],
                                    lesson['time'], lesson['subject'], 
                                    lesson['classroom'], status
                                ))

                day_lessons.sort(key=lambda x: x[2])
                for lesson_data in day_lessons:
                    tree.insert("", tk.END, values=lesson_data)

            total_label.config(text=f"Всего уроков: {total_lessons}")
            sub_label.config(text=f"Замен: {total_subs}")

            if total_lessons == 0:
                tree.insert("", tk.END, values=("Учитель не найден", "", "", "", "", "", ""))

        ttk.Button(win, text="🔄 Показать расписание", command=update_schedule).pack(pady=5)

        if self.all_teachers:
            teacher_combo.set(self.all_teachers[0])
            win.after(100, update_schedule)

    def search_dialog(self):
        """Диалог поиска"""
        win = tk.Toplevel(self.root)
        win.title("Поиск")
        win.geometry("700x600")

        search_frame = ttk.Frame(win)
        search_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(search_frame, text="Найти:", font=("Arial", 11, "bold")).pack(side=tk.LEFT, padx=5)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30, font=("Arial", 10))
        search_entry.pack(side=tk.LEFT, padx=5)

        type_frame = ttk.Frame(win)
        type_frame.pack(fill=tk.X, padx=10, pady=5)
        
        search_type = tk.StringVar(value="all")
        ttk.Radiobutton(type_frame, text="🔍 Везде", variable=search_type, 
                       value="all").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(type_frame, text="👤 Учитель", variable=search_type, 
                       value="teacher").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(type_frame, text="📚 Предмет", variable=search_type, 
                       value="subject").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(type_frame, text="🏫 Кабинет", variable=search_type, 
                       value="room").pack(side=tk.LEFT, padx=10)

        result_frame = ttk.LabelFrame(win, text="Результаты поиска")
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        result_text = tk.Text(result_frame, wrap=tk.WORD, font=("Arial", 10))
        scroll = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=result_text.yview)
        result_text.configure(yscrollcommand=scroll.set)
        
        result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        def do_search():
            query = search_var.get().strip().lower()
            if not query:
                messagebox.showwarning("Ошибка", "Введите текст для поиска")
                return
            
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, f"🔍 РЕЗУЛЬТАТЫ ПОИСКА: '{query}'\n", "header")
            result_text.insert(tk.END, "="*70 + "\n\n")
            
            found = False
            search_all = search_type.get() == "all"
            
            for day in self.days:
                for class_num in self.classes:
                    if day in self.weekly_schedule and class_num in self.weekly_schedule[day]:
                        for lesson in self.weekly_schedule[day][class_num]:
                            match = False
                            if search_all:
                                match = (query in lesson['teacher'].lower() or 
                                        query in lesson['subject'].lower() or 
                                        query in lesson['classroom'].lower())
                            elif search_type.get() == "teacher":
                                match = query in lesson['teacher'].lower()
                            elif search_type.get() == "subject":
                                match = query in lesson['subject'].lower()
                            elif search_type.get() == "room":
                                match = query in lesson['classroom'].lower()
                            
                            if match:
                                found = True
                                result_text.insert(tk.END, f"📅 {day}, {class_num} класс\n")
                                result_text.insert(tk.END, f"   Урок {lesson['number']}: {lesson['time']}\n")
                                result_text.insert(tk.END, f"   📚 {lesson['subject']}\n")
                                result_text.insert(tk.END, f"   👤 {lesson['teacher']}\n")
                                result_text.insert(tk.END, f"   🏫 {lesson['classroom']}\n")
                                if lesson['substituted'] or lesson['room_changed']:
                                    status = []
                                    if lesson['substituted']:
                                        status.append("замена учителя")
                                    if lesson['room_changed']:
                                        status.append("смена кабинета")
                                    result_text.insert(tk.END, f"   ⚠️ {', '.join(status)}\n")
                                result_text.insert(tk.END, "-"*50 + "\n\n")
            
            if not found:
                result_text.insert(tk.END, "❌ Ничего не найдено\n")

        btn_frame = ttk.Frame(win)
        btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(btn_frame, text="🔍 Найти", command=do_search, width=20).pack()

        search_entry.bind('<Return>', lambda e: do_search())
        
        result_text.tag_configure("header", font=("Arial", 12, "bold"), foreground="blue")

    def show_class_stats(self):
        """Статистика по классам"""
        win = tk.Toplevel(self.root)
        win.title("Статистика по классам")
        win.geometry("800x500")

        columns = ("Класс", "Уровень", "Всего уроков", "Замен учит.", "Смен каб.", "Всего измен.")
        tree = ttk.Treeview(win, columns=columns, show="headings", height=20)
        
        col_widths = [60, 120, 80, 80, 80, 80]
        for col, width in zip(columns, col_widths):
            tree.heading(col, text=col)
            tree.column(col, width=width, anchor="center")
        
        scroll = ttk.Scrollbar(win, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        total_lessons_all = 0
        total_subs_all = 0
        total_rooms_all = 0

        for class_num in self.classes:
            if class_num <= 4:
                level = "Начальная школа"
            elif class_num <= 9:
                level = "Средняя школа"
            else:
                level = "Старшая школа"
            
            lessons_count = 0
            teacher_subs = 0
            room_changes = 0
            
            for day in self.days:
                if day in self.weekly_schedule and class_num in self.weekly_schedule[day]:
                    for lesson in self.weekly_schedule[day][class_num]:
                        lessons_count += 1
                        if lesson['substituted']:
                            teacher_subs += 1
                        if lesson['room_changed']:
                            room_changes += 1
            
            total_changes = teacher_subs + room_changes
            
            total_lessons_all += lessons_count
            total_subs_all += teacher_subs
            total_rooms_all += room_changes
            
            tree.insert("", tk.END, values=(
                f"{class_num} класс", level, lessons_count,
                teacher_subs, room_changes, total_changes
            ))

        tree.insert("", tk.END, values=(
            "ВСЕГО", f"{len(self.classes)} классов", 
            total_lessons_all, total_subs_all, total_rooms_all,
            total_subs_all + total_rooms_all
        ))
        
        tree.tag_configure("total", background="#E0E0E0", font=("Arial", 10, "bold"))
        last_item = tree.get_children()[-1]
        tree.item(last_item, tags=("total",))
    
    def show_all_events(self):
        """Показывает все мероприятия на неделю по классам"""
        win = tk.Toplevel(self.root)
        win.title("Все мероприятия на неделю")
        win.geometry("900x700")

        main_notebook = ttk.Notebook(win)
        main_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        for day in self.days:
            day_frame = ttk.Frame(main_notebook)
            main_notebook.add(day_frame, text=day)

            class_notebook = ttk.Notebook(day_frame)
            class_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

            class_frames = {}
            for class_num in self.classes:
                class_frame = ttk.Frame(class_notebook)
                class_notebook.add(class_frame, text=f"{class_num} класс")
                class_frames[class_num] = class_frame

            has_any_events = False
            for class_num in self.classes:
                frame = class_frames[class_num]

                if day in self.events and class_num in self.events[day] and self.events[day][class_num]:
                    has_any_events = True
                    
                    text = tk.Text(frame, wrap=tk.WORD, font=("Arial", 10))
                    scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text.yview)
                    text.configure(yscrollcommand=scroll.set)
                    
                    text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                    scroll.pack(side=tk.RIGHT, fill=tk.Y)
                    
                    for event in self.events[day][class_num]:
                        color = self.event_colors.get(event['type'], "#A9A9A9")
                        
                        text.insert(tk.END, f"🕒 {event['time']}\n", "time")
                        text.insert(tk.END, f"📌 {event['type']}: {event['name']}\n", "name")
                        text.insert(tk.END, f"👤 Ответственный: {event['teacher']}\n", "teacher")
                        text.insert(tk.END, f"📍 Место: {event['place']}\n", "place")
                        text.insert(tk.END, f"👥 Участники: {event['participants']}\n", "participants")
                        if event['description']:
                            text.insert(tk.END, f"📝 {event['description']}\n", "desc")
                        text.insert(tk.END, "-"*50 + "\n\n", "separator")
                    
                    text.tag_configure("time", font=("Arial", 10, "bold"), foreground="blue")
                    text.tag_configure("name", font=("Arial", 11, "bold"), foreground="darkblue")
                    text.tag_configure("teacher", font=("Arial", 10), foreground="green")
                    text.tag_configure("place", font=("Arial", 10), foreground="brown")
                    text.tag_configure("participants", font=("Arial", 10), foreground="gray")
                    text.tag_configure("desc", font=("Arial", 9, "italic"), foreground="gray")
                    text.tag_configure("separator", font=("Arial", 8))
                    text.config(state=tk.DISABLED)
                else:
                    label = ttk.Label(frame, text="Нет мероприятий для этого класса", 
                                     font=("Arial", 12))
                    label.pack(expand=True)
    
    def get_current_schedule(self):
        """Возвращает текущее расписание для выбранного дня и класса"""
        try:
            day = self.sub_day.get()
            class_num = int(self.sub_class.get())
            if day and class_num and day in self.weekly_schedule:
                return self.weekly_schedule[day][class_num]
        except:
            pass
        return None
    
    def get_current_schedule_for_room(self):
        """Возвращает текущее расписание для выбранного дня и класса (для замены кабинета)"""
        try:
            day = self.room_day.get()
            class_num = int(self.room_class.get())
            if day and class_num and day in self.weekly_schedule:
                return self.weekly_schedule[day][class_num]
        except:
            pass
        return None
    
    def update_teachers_for_sub(self, event=None):
        """Обновляет список учителей для замены"""
        try:
            day = self.sub_day.get()
            class_num = int(self.sub_class.get())
            lesson_num = int(self.sub_lesson.get())
            
            schedule = self.get_current_schedule()
            if not schedule:
                return
            
            for lesson in schedule:
                if lesson['number'] == lesson_num:
                    subject = lesson['subject']
                    current = lesson['teacher']
                    
                    teachers = self.teachers.get(subject, [])
                    available = [t for t in teachers if t != current]
                    
                    if available:
                        self.sub_teacher['values'] = available
                        self.sub_teacher.set(available[0])
                    else:
                        self.sub_teacher['values'] = ["Нет замены"]
                        self.sub_teacher.set("Нет замены")
                    break
        except:
            pass
    
    def manual_substitute(self):
        """Ручная замена учителя"""
        try:
            day = self.sub_day.get()
            class_num = int(self.sub_class.get())
            lesson_num = int(self.sub_lesson.get())
            new_teacher = self.sub_teacher.get()
            
            if not day or not class_num:
                messagebox.showwarning("Ошибка", "Выберите день и класс")
                return
            
            if not new_teacher or new_teacher == "Нет замены":
                messagebox.showwarning("Ошибка", "Выберите учителя")
                return
            
            schedule = self.get_current_schedule()
            if not schedule:
                return
            
            for lesson in schedule:
                if lesson['number'] == lesson_num:
                    if lesson['substituted']:
                        messagebox.showinfo("Инфо", "Уже есть замена")
                        return

                    sub_info = {
                        'datetime': datetime.now().strftime("%Y-%m-%d %H:%M"),
                        'type': 'teacher',
                        'day': day,
                        'class': class_num,
                        'lesson': lesson_num,
                        'subject': lesson['subject'],
                        'from': lesson['teacher'],
                        'to': new_teacher
                    }
                    self.substitutions.append(sub_info)

                    lesson['original_teacher'] = lesson['teacher']
                    lesson['teacher'] = new_teacher
                    lesson['substituted'] = True

                    self.update_day_display(day, class_num)
                    self.status.config(text=f"✅ Замена учителя: {day}, {class_num} класс, урок {lesson_num}")
                    return
                    
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
    
    def manual_room_change(self):
        """Ручная замена кабинета"""
        try:
            day = self.room_day.get()
            class_num = int(self.room_class.get())
            lesson_num = int(self.room_lesson.get())
            new_room = self.room_entry.get().strip()
            
            if not day or not class_num:
                messagebox.showwarning("Ошибка", "Выберите день и класс")
                return
            
            if not new_room:
                messagebox.showwarning("Ошибка", "Введите новый кабинет")
                return
            
            schedule = self.get_current_schedule_for_room()
            if not schedule:
                return
            
            for lesson in schedule:
                if lesson['number'] == lesson_num:
                    room_info = {
                        'datetime': datetime.now().strftime("%Y-%m-%d %H:%M"),
                        'type': 'room',
                        'day': day,
                        'class': class_num,
                        'lesson': lesson_num,
                        'subject': lesson['subject'],
                        'from': lesson['classroom'],
                        'to': new_room
                    }
                    self.classroom_changes.append(room_info)

                    lesson['original_classroom'] = lesson['classroom']
                    lesson['classroom'] = new_room
                    lesson['room_changed'] = True

                    self.update_day_display(day, class_num)
                    self.status.config(text=f"✅ Смена кабинета: {day}, {class_num} класс, урок {lesson_num} -> {new_room}")
                    return
                    
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
    
    def update_day_display(self, day, class_num):
        """Обновляет отображение для конкретного дня и класса"""
        if day not in self.weekly_schedule or class_num not in self.weekly_schedule[day]:
            return
        
        tree = self.day_trees[day][class_num]
        tree.delete(*tree.get_children())
        
        for lesson in self.weekly_schedule[day][class_num]:
            status = ""
            status_color = "white"
            
            if lesson['substituted'] and lesson['room_changed']:
                status = "✅👥"
                status_color = "#DDA0DD"  
            elif lesson['substituted']:
                status = "✅👤"
                status_color = "#FFE4B5"  
            elif lesson['room_changed']:
                status = "✅🏫"
                status_color = "#E0FFFF"  
            else:
                status = "✓"
                status_color = "#FFFFFF"  
            
            teacher = lesson['teacher']
            if lesson['substituted']:
                teacher = f"🔁 {teacher}"
            
            classroom = lesson['classroom']
            if lesson['room_changed']:
                classroom = f"🏫 {classroom}"
            
            item = tree.insert("", tk.END, values=(
                lesson['number'],
                lesson['time'],
                lesson['subject'],
                teacher,
                classroom,
                status
            ))

            subject_color = self.subject_colors.get(lesson['subject'], "white")
            tree.tag_configure(lesson['subject'], background=subject_color)

            if status != "✓":
                tree.tag_configure(f"status_{lesson['number']}", background=status_color)
                tree.item(item, tags=(lesson['subject'], f"status_{lesson['number']}"))
            else:
                tree.item(item, tags=(lesson['subject'],))
    
    def show_teacher_load(self):
        """Показывает нагрузку учителей"""
        win = tk.Toplevel(self.root)
        win.title("Статистика")
        win.geometry("600x500")
        
        text = tk.Text(win, wrap=tk.WORD, font=("Courier", 10))
        scroll = ttk.Scrollbar(win, orient=tk.VERTICAL, command=text.yview)
        text.configure(yscrollcommand=scroll.set)
        
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        text.insert(tk.END, "📊 НАГРУЗКА УЧИТЕЛЕЙ\n", "header")
        text.insert(tk.END, "="*50 + "\n\n")
        
        if not self.teacher_load:
            text.insert(tk.END, "Нет данных\n")
        else:
            sorted_load = sorted(self.teacher_load.items(), key=lambda x: x[1], reverse=True)
            for teacher, hours in sorted_load:
                bar = "█" * min(hours, 30)
                text.insert(tk.END, f"{teacher:<25} {hours:2d} ч {bar}\n")
        
        text.insert(tk.END, "\n\n📊 КОЛИЧЕСТВО ЗАМЕН\n", "header")
        text.insert(tk.END, "="*50 + "\n\n")
        text.insert(tk.END, f"Замен учителей: {len(self.substitutions)}\n")
        text.insert(tk.END, f"Смен кабинетов: {len(self.classroom_changes)}\n")
        
        text.tag_configure("header", font=("Arial", 12, "bold"))
        text.config(state=tk.DISABLED)
    
    def show_substitutions(self):
        """Показывает журнал замен"""
        if not self.substitutions and not self.classroom_changes:
            messagebox.showinfo("Журнал", "Журнал замен пуст")
            return
        
        win = tk.Toplevel(self.root)
        win.title("Журнал замен")
        win.geometry("800x600")
        
        notebook = ttk.Notebook(win)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        teacher_frame = ttk.Frame(notebook)
        notebook.add(teacher_frame, text="Замены учителей")
        
        teacher_text = tk.Text(teacher_frame, wrap=tk.WORD, font=("Courier", 10))
        teacher_scroll = ttk.Scrollbar(teacher_frame, orient=tk.VERTICAL, command=teacher_text.yview)
        teacher_text.configure(yscrollcommand=teacher_scroll.set)
        
        teacher_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        teacher_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        if self.substitutions:
            for i, sub in enumerate(self.substitutions, 1):
                teacher_text.insert(tk.END, f"ЗАМЕНА УЧИТЕЛЯ #{i}\n", "header")
                teacher_text.insert(tk.END, f"  Время: {sub['datetime']}\n")
                teacher_text.insert(tk.END, f"  День: {sub['day']}, класс {sub['class']}\n")
                teacher_text.insert(tk.END, f"  Урок {sub['lesson']}: {sub['subject']}\n")
                teacher_text.insert(tk.END, f"  Было: {sub['from']}\n")
                teacher_text.insert(tk.END, f"  Стало: {sub['to']}\n")
                teacher_text.insert(tk.END, "-"*40 + "\n\n")
        else:
            teacher_text.insert(tk.END, "Нет замен учителей\n")

        room_frame = ttk.Frame(notebook)
        notebook.add(room_frame, text="Смены кабинетов")
        
        room_text = tk.Text(room_frame, wrap=tk.WORD, font=("Courier", 10))
        room_scroll = ttk.Scrollbar(room_frame, orient=tk.VERTICAL, command=room_text.yview)
        room_text.configure(yscrollcommand=room_scroll.set)
        
        room_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        room_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        if self.classroom_changes:
            for i, ch in enumerate(self.classroom_changes, 1):
                room_text.insert(tk.END, f"СМЕНА КАБИНЕТА #{i}\n", "header")
                room_text.insert(tk.END, f"  Время: {ch['datetime']}\n")
                room_text.insert(tk.END, f"  День: {ch['day']}, класс {ch['class']}\n")
                room_text.insert(tk.END, f"  Урок {ch['lesson']}: {ch['subject']}\n")
                room_text.insert(tk.END, f"  Был: {ch['from']}\n")
                room_text.insert(tk.END, f"  Стал: {ch['to']}\n")
                room_text.insert(tk.END, "-"*40 + "\n\n")
        else:
            room_text.insert(tk.END, "Нет смен кабинетов\n")
        
        teacher_text.tag_configure("header", font=("Arial", 11, "bold"), foreground="blue")
        room_text.tag_configure("header", font=("Arial", 11, "bold"), foreground="blue")
        teacher_text.config(state=tk.DISABLED)
        room_text.config(state=tk.DISABLED)
    
    def save_to_txt(self):
        """Сохраняет расписание в форматированный текстовый файл"""
        if not self.weekly_schedule:
            messagebox.showwarning("Ошибка", "Сначала создайте расписание")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("="*100 + "\n")
                f.write(" " * 40 + "ШКОЛЬНОЕ РАСПИСАНИЕ\n")
                f.write("="*100 + "\n\n")
                f.write(f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
                f.write(f"Количество уроков: от {min(self.lessons_count.values())} до {max(self.lessons_count.values())}\n\n")

                f.write("СТАТУСЫ УРОКОВ:\n")
                f.write("  ✓ - Обычный урок\n")
                f.write("  ✅👤 - Замена учителя\n")
                f.write("  ✅🏫 - Смена кабинета\n")
                f.write("  ✅👥 - Замена учителя и кабинета\n\n")

                for day in self.days:
                    f.write("\n" + "="*100 + "\n")
                    f.write(f"{day}\n".center(100) + "\n")
                    f.write("="*100 + "\n\n")
 
                    for class_num in self.classes:
                        if day in self.weekly_schedule and class_num in self.weekly_schedule[day]:
                            if class_num <= 4:
                                level = "Начальная школа"
                            elif class_num <= 9:
                                level = "Средняя школа"
                            else:
                                level = "Старшая школа"
                            
                            f.write(f"\n📚 {class_num} КЛАСС ({level})\n")
                            f.write("-"*90 + "\n")
                            f.write(f"{'№':<4} {'Время':<15} {'Предмет':<25} {'Учитель':<25} {'Кабинет':<10} Статус\n")
                            f.write("-"*90 + "\n")
                            
                            for lesson in self.weekly_schedule[day][class_num]:
                                status = ""
                                if lesson.get('substituted', False) and lesson.get('room_changed', False):
                                    status = "✅👥"
                                elif lesson.get('substituted', False):
                                    status = "✅👤"
                                elif lesson.get('room_changed', False):
                                    status = "✅🏫"
                                else:
                                    status = "✓"
                                
                                teacher = lesson['teacher']
                                if lesson.get('substituted', False):
                                    teacher = f"🔁 {teacher}"
                                
                                classroom = lesson['classroom']
                                if lesson.get('room_changed', False):
                                    classroom = f"🏫 {classroom}"
                                
                                f.write(f"{lesson['number']:<4} "
                                       f"{lesson['time']:<15} "
                                       f"{lesson['subject']:<25} "
                                       f"{teacher:<25} "
                                       f"{classroom:<10} "
                                       f"{status}\n")

                    has_events = False
                    for class_num in self.classes:
                        if day in self.events and class_num in self.events[day] and self.events[day][class_num]:
                            if not has_events:
                                f.write(f"\n🎉 МЕРОПРИЯТИЯ НА {day.upper()}\n")
                                f.write("-"*90 + "\n")
                                has_events = True
                            
                            f.write(f"\n📌 {class_num} КЛАСС:\n")
                            for event in self.events[day][class_num]:
                                f.write(f"   🕒 {event['time']} [{event['type']}] {event['name']}\n")
                                f.write(f"      👤 {event['teacher']}\n")
                                f.write(f"      📍 {event['place']}\n")
                                f.write(f"      👥 {event['participants']}\n")
                                if event.get('description'):
                                    f.write(f"      📝 {event['description']}\n")
                                f.write("   " + "-"*40 + "\n")
                    
                    if has_events:
                        f.write("\n" + "="*100 + "\n")

                f.write("\n\n" + "="*100 + "\n")
                f.write(" " * 40 + "СТАТИСТИКА УЧИТЕЛЕЙ\n")
                f.write("="*100 + "\n\n")
                
                if self.teacher_load:
                    sorted_load = sorted(self.teacher_load.items(), key=lambda x: x[1], reverse=True)
                    
                    f.write(f"{'Учитель':<30} {'Количество уроков':<20}\n")
                    f.write("-"*50 + "\n")
                    for teacher, count in sorted_load:
                        bar = "█" * min(count, 30)
                        f.write(f"{teacher:<30} {count:<20} {bar}\n")
                    
                    f.write("\n" + "-"*50 + "\n")
                    f.write(f"Всего учителей: {len(self.teacher_load)}\n")
                    f.write(f"Всего уроков: {sum(self.teacher_load.values())}\n")
                    avg = sum(self.teacher_load.values()) / len(self.teacher_load) if self.teacher_load else 0
                    f.write(f"Средняя нагрузка: {avg:.1f} уроков\n")

                f.write("\n\n" + "="*100 + "\n")
                f.write(" " * 40 + "СТАТИСТИКА ЗАМЕН\n")
                f.write("="*100 + "\n\n")
                f.write(f"Замен учителей: {len(self.substitutions)}\n")
                f.write(f"Смен кабинетов: {len(self.classroom_changes)}\n")
                
                if self.substitutions:
                    f.write("\nПоследние замены учителей:\n")
                    for sub in self.substitutions[-5:]:
                        f.write(f"  • {sub['day']} {sub['class']} кл, урок {sub['lesson']}: {sub['from']} -> {sub['to']}\n")
                
                if self.classroom_changes:
                    f.write("\nПоследние смены кабинетов:\n")
                    for ch in self.classroom_changes[-5:]:
                        f.write(f"  • {ch['day']} {ch['class']} кл, урок {ch['lesson']}: {ch['from']} -> {ch['to']}\n")
            
            self.status.config(text=f"✅ Расписание сохранено в TXT: {filename}")
            messagebox.showinfo("Успех", f"Расписание сохранено в TXT файл:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить TXT: {e}")
    
    def save_to_excel(self):
        """Сохраняет расписание в Excel файл"""
        if not EXCEL_AVAILABLE:
            messagebox.showwarning("Внимание", 
                                  "Библиотека openpyxl не установлена.\n"
                                  "Установите: pip install openpyxl")
            return
        
        if not self.weekly_schedule:
            messagebox.showwarning("Ошибка", "Сначала создайте расписание")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()

            wb.remove(wb.active)

            header_font = Font(name='Arial', size=12, bold=True)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font_white = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            
            day_font = Font(name='Arial', size=14, bold=True)
            day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            class_font = Font(name='Arial', size=11, bold=True)
            class_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            event_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            event_class_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")

            sub_color = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  
            room_color = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  
            both_color = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")  
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for day in self.days:
                ws = wb.create_sheet(title=day[:8])

                ws.merge_cells('A1:F1')
                cell = ws['A1']
                cell.value = f"РАСПИСАНИЕ - {day}"
                cell.font = day_font
                cell.fill = day_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
 
                headers = ['Класс', '№', 'Время', 'Предмет', 'Учитель', 'Кабинет']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=col)
                    cell.value = header
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                row = 3
                for class_num in self.classes:
                    if day in self.weekly_schedule and class_num in self.weekly_schedule[day]:
                        ws.merge_cells(start_row=row, start_column=1, end_row=row+5, end_column=1)
                        class_cell = ws.cell(row=row, column=1)
                        class_cell.value = f"{class_num} класс"
                        class_cell.font = class_font
                        class_cell.fill = class_fill
                        class_cell.alignment = Alignment(horizontal='center', vertical='center')
                        class_cell.border = thin_border

                        for lesson in self.weekly_schedule[day][class_num]:
                            ws.cell(row=row, column=2).value = lesson['number']
                            ws.cell(row=row, column=3).value = lesson['time']
                            ws.cell(row=row, column=4).value = lesson['subject']
                            
                            teacher = lesson['teacher']
                            if lesson.get('substituted', False):
                                teacher = f"🔁 {teacher}"
                            ws.cell(row=row, column=5).value = teacher
                            
                            classroom = lesson['classroom']
                            if lesson.get('room_changed', False):
                                classroom = f"🏫 {classroom}"
                            ws.cell(row=row, column=6).value = classroom

                            if lesson.get('substituted', False) and lesson.get('room_changed', False):
                                fill = both_color
                            elif lesson.get('substituted', False):
                                fill = sub_color
                            elif lesson.get('room_changed', False):
                                fill = room_color
                            else:
                                fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                            
                            for col in range(2, 7):
                                cell = ws.cell(row=row, column=col)
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center')
                                cell.fill = fill
                            
                            row += 1
                        
                        row += 1

                has_events = False
                for class_num in self.classes:
                    if day in self.events and class_num in self.events[day] and self.events[day][class_num]:
                        if not has_events:
                            row += 1
                            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                            event_header = ws.cell(row=row, column=1)
                            event_header.value = f"🎉 МЕРОПРИЯТИЯ - {day}"
                            event_header.font = Font(bold=True, size=12)
                            event_header.fill = event_fill
                            event_header.alignment = Alignment(horizontal='center')
                            row += 1
                            has_events = True

                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                        class_event_header = ws.cell(row=row, column=1)
                        class_event_header.value = f"📌 {class_num} КЛАСС"
                        class_event_header.font = Font(bold=True)
                        class_event_header.fill = event_class_fill
                        class_event_header.alignment = Alignment(horizontal='center')
                        row += 1
                        
                        for event in self.events[day][class_num]:
                            ws.cell(row=row, column=1).value = "🕒"
                            ws.cell(row=row, column=2).value = event['time']
                            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
                            event_cell = ws.cell(row=row, column=3)
                            event_cell.value = f"{event['type']}: {event['name']}"
                            event_cell.font = Font(bold=True)
                            row += 1
                            
                            ws.cell(row=row, column=1).value = "👤"
                            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                            ws.cell(row=row, column=2).value = f"Ответственный: {event['teacher']}"
                            row += 1
                            
                            ws.cell(row=row, column=1).value = "📍"
                            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                            ws.cell(row=row, column=2).value = f"Место: {event['place']}"
                            row += 1
                            
                            if event.get('participants'):
                                ws.cell(row=row, column=1).value = "👥"
                                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                                ws.cell(row=row, column=2).value = f"Участники: {event['participants']}"
                                row += 1
                            
                            if event.get('description'):
                                ws.cell(row=row, column=1).value = "📝"
                                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                                ws.cell(row=row, column=2).value = event['description']
                                row += 1
                            
                            row += 1

                for col in range(1, 7):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 25

            stats_ws = wb.create_sheet(title="Статистика", index=0)
            
            stats_ws.merge_cells('A1:C1')
            title_cell = stats_ws['A1']
            title_cell.value = "СТАТИСТИКА УЧИТЕЛЕЙ"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            stats_ws['A3'].value = "Учитель"
            stats_ws['B3'].value = "Уроков"
            stats_ws['C3'].value = "График"
            
            for col in ['A3', 'B3', 'C3']:
                stats_ws[col].font = Font(bold=True)
                stats_ws[col].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                stats_ws[col].font = Font(bold=True, color="FFFFFF")
            
            row = 4
            if self.teacher_load:
                sorted_load = sorted(self.teacher_load.items(), key=lambda x: x[1], reverse=True)
                for teacher, count in sorted_load:
                    stats_ws[f'A{row}'].value = teacher
                    stats_ws[f'B{row}'].value = count
                    stats_ws[f'C{row}'].value = "█" * min(count, 30)
                    row += 1
            
            stats_ws.column_dimensions['A'].width = 30
            stats_ws.column_dimensions['B'].width = 10
            stats_ws.column_dimensions['C'].width = 35

            wb.save(filename)
            
            self.status.config(text=f"✅ Расписание сохранено в Excel: {filename}")
            messagebox.showinfo("Успех", f"Расписание сохранено в Excel файл:\n{filename}")
            
        except ImportError:
            messagebox.showwarning("Внимание", 
                                  "Библиотека openpyxl не установлена.\n"
                                  "Установите: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить Excel: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ScheduleApp(root)
    root.mainloop()
